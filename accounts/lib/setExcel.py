from django.http import HttpResponse
from cost.models import Category, Shop, Cost
from sales.models import Sales
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill

"""領収書出力"""
def makeCostExe(request, year, month):
    excel_cost_list = []
    food_shop_list = []
    other_shop_list = []
    excel_day_column = 1
    excel_price_column = 2
    fill = PatternFill(fill_type='solid', fgColor='d5e6fc')
    border = Border(
        left=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        right=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        top=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        bottom=Side(
            border_style="thin",
            color="A7A7A7"
        )
    )

    """食材リスト作成"""
    Food_shop_data = Shop.objects.filter(flag=0, category__category_name="食材")
    for f in Food_shop_data:
        food_shop_list.append([f.category.category_name, f.shop_name])

    """その他リスト作成"""
    Other_shop_data = Shop.objects.filter(flag=0).exclude(category__category_name="食材")
    for o in Other_shop_data:
        other_shop_list.append([o.category.category_name, o.shop_name])

    """該当月の経費(領収証)データ作成"""
    Cost_month_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__flag=0)
    for c in Cost_month_data:
        date = c.date.strftime("%m/%d")
        excel_cost_list.append([c.shop.category.category_name, c.shop.shop_name, date, c.price])

    """Excel出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '現金仕入(領収証){0}年{1}月度'.format(year, month)
    """食材"""
    for data in food_shop_list:
        ws.cell(row=1, column=1, value=data[0])
        ws.cell(row=2, column=excel_day_column, value="日付")
        sn = ws.cell(row=2, column=excel_price_column, value=data[1])
        sn.fill = fill
        sn.border = border
        excel_cost_row = 3
        for category, shop, day, price in excel_cost_list:
            if data[0] == category and data[1] == shop:
                ws.cell(row=excel_cost_row, column=excel_day_column, value=day)
                ws.cell(row=excel_cost_row, column=excel_price_column, value=price)
                excel_cost_row += 1
        min_address = ws.cell(row=3,column=excel_price_column).coordinate
        max_address = ws.cell(row=excel_cost_row,column=excel_price_column).coordinate
        ws.cell(row=excel_cost_row + 1, column=excel_day_column, value="合計")
        ws.cell(row=excel_cost_row + 1, column=excel_price_column, value='=SUM({}:{})'.format(min_address, max_address))
        excel_day_column += 2
        excel_price_column += 2
    """その他現金経費"""
    max_row = ws.max_row + 3
    excel_day_column = 1
    excel_price_column = 2
    for data in other_shop_list:
        ws.cell(row=max_row, column=1, value="その他 現金経費")
        ws.cell(row=max_row + 1, column=excel_day_column, value=data[0])
        ws.cell(row=max_row + 2, column=excel_day_column, value="日付")
        sn = ws.cell(row=max_row + 2, column=excel_price_column, value=data[1])
        sn.fill = fill
        sn.border = border
        excel_cost_row = max_row + 3
        for category, shop, day, price in excel_cost_list:
            if data[0] == category and data[1] == shop:
                ws.cell(row=excel_cost_row, column=excel_day_column, value=day)
                ws.cell(row=excel_cost_row, column=excel_price_column, value=price)
                excel_cost_row += 1
        min_address = ws.cell(row=max_row + 3,column=excel_price_column).coordinate
        max_address = ws.cell(row=excel_cost_row,column=excel_price_column).coordinate
        ws.cell(row=excel_cost_row + 1, column=excel_day_column, value="合計")
        ws.cell(row=excel_cost_row + 1, column=excel_price_column, value='=SUM({}:{})'.format(min_address, max_address))
        excel_day_column += 2
        excel_price_column += 2

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'cost_{0}_{1}.xlsx'.format(year, month)

    wb.save(response)
    return response

"""納品書出力"""
def makeDeliveryExe(request, year, month):
    excel_delivery_list = []
    food_delivery_shop_list = []
    drink_delivery_shop_list = []
    other_delivery_shop_list = []
    excel_delivery_day_column = 1
    excel_delivery_price_column = 2
    fill = PatternFill(fill_type='solid', fgColor='62daa8')
    border = Border(
        left=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        right=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        top=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        bottom=Side(
            border_style="thin",
            color="A7A7A7"
        )
    )
    """食材リスト作成"""
    Food_delivery_shop_data = Shop.objects.filter(flag=1, category__category_name="食材")
    for fd in Food_delivery_shop_data:
        food_delivery_shop_list.append([fd.category.category_name, fd.shop_name])

    """ドリンクリスト作成"""
    Drink_delivery_shop_data = Shop.objects.filter(flag=1, category__category_name="ドリンク")
    for dd in Drink_delivery_shop_data:
        drink_delivery_shop_list.append([dd.category.category_name, dd.shop_name])

    """その他リスト作成"""
    Other_delivery_shop_data = Shop.objects.filter(flag=1, category__category_name="その他")
    for od in Other_delivery_shop_data:
        other_delivery_shop_list.append([od.category.category_name, od.shop_name])

    """該当月の経費(納品書)データ作成"""
    Delivery_month_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__flag=1)
    for d in Delivery_month_data:
        date = d.date.strftime("%m/%d")
        excel_delivery_list.append([d.shop.category.category_name, d.shop.shop_name, date, d.price])

    """Excel出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '納品書{0}年{1}月度'.format(year, month)
    """食材"""
    for data in food_delivery_shop_list:
        ws.cell(row=1, column=1, value=data[0])
        ws.cell(row=2, column=excel_delivery_day_column, value="日付")
        sn = ws.cell(row=2, column=excel_delivery_price_column, value=data[1])
        sn.fill = fill
        sn.border = border
        excel_delivery_cost_row = 3
        for category, shop, day, price in excel_delivery_list:
            if data[0] == category and data[1] == shop:
                ws.cell(row=excel_delivery_cost_row, column=excel_delivery_day_column, value=day)
                ws.cell(row=excel_delivery_cost_row, column=excel_delivery_price_column, value=price)
                excel_delivery_cost_row += 1
        min_address = ws.cell(row=3,column=excel_delivery_price_column).coordinate
        max_address = ws.cell(row=excel_delivery_cost_row,column=excel_delivery_price_column).coordinate
        ws.cell(row=excel_delivery_cost_row + 1, column=excel_delivery_day_column, value="合計")
        ws.cell(row=excel_delivery_cost_row + 1, column=excel_delivery_price_column, value='=SUM({}:{})'.format(min_address, max_address))
        excel_delivery_day_column += 2
        excel_delivery_price_column += 2
    """ドリンク"""
    delivery_max_row = ws.max_row + 3
    excel_delivery_day_column = 1
    excel_delivery_price_column = 2
    for data in drink_delivery_shop_list:
        ws.cell(row=delivery_max_row, column=1, value=data[0])
        ws.cell(row=delivery_max_row + 1, column=excel_delivery_day_column, value="日付")
        sn = ws.cell(row=delivery_max_row + 1, column=excel_delivery_price_column, value=data[1])
        sn.fill = fill
        sn.border = border
        excel_delivery_row = delivery_max_row + 2
        for category, shop, day, price in excel_delivery_list:
            if data[0] == category and data[1] == shop:
                ws.cell(row=excel_delivery_row, column=excel_delivery_day_column, value=day)
                ws.cell(row=excel_delivery_row, column=excel_delivery_price_column, value=price)
                excel_delivery_row += 1
        min_address = ws.cell(row=delivery_max_row + 2,column=excel_delivery_price_column).coordinate
        max_address = ws.cell(row=excel_delivery_row,column=excel_delivery_price_column).coordinate
        ws.cell(row=excel_delivery_row + 1, column=excel_delivery_day_column, value="合計")
        ws.cell(row=excel_delivery_row + 1, column=excel_delivery_price_column, value='=SUM({}:{})'.format(min_address, max_address))
        excel_delivery_day_column += 2
        excel_delivery_price_column += 2
    """その他"""
    delivery_max_row = ws.max_row + 3
    excel_delivery_day_column = 1
    excel_delivery_price_column = 2
    for data in other_delivery_shop_list:
        ws.cell(row=delivery_max_row, column=1, value=data[0])
        ws.cell(row=delivery_max_row + 1, column=excel_delivery_day_column, value="日付")
        sn = ws.cell(row=delivery_max_row + 1, column=excel_delivery_price_column, value=data[1])
        sn.fill = fill
        sn.border = border
        excel_delivery_row = delivery_max_row + 2
        for category, shop, day, price in excel_delivery_list:
            if data[0] == category and data[1] == shop:
                ws.cell(row=excel_delivery_row, column=excel_delivery_day_column, value=day)
                ws.cell(row=excel_delivery_row, column=excel_delivery_price_column, value=price)
                excel_delivery_row += 1
        min_address = ws.cell(row=delivery_max_row + 2,column=excel_delivery_price_column).coordinate
        max_address = ws.cell(row=excel_delivery_row,column=excel_delivery_price_column).coordinate
        ws.cell(row=excel_delivery_row + 1, column=excel_delivery_day_column, value="合計")
        ws.cell(row=excel_delivery_row + 1, column=excel_delivery_price_column, value='=SUM({}:{})'.format(min_address, max_address))
        excel_delivery_day_column += 2
        excel_delivery_price_column += 2

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'delivery_{0}_{1}.xlsx'.format(year, month)

    wb.save(response)
    return response

"""日報出力"""
def makeSalesExe(request, year, month):
    excel_sales_list = [] # 該当月の日報データを格納
    excel_sales_row = 2
    excel_sales_sales_total = 0 # 該当月の売上合計金額
    excel_sales_dinner_total = 0 # 該当月のディナー売上合計金額
    excel_sales_lunch_total = 0 # 該当月のランチ売上合計金額
    excel_sales_food_total = 0 # 該当月のフード売上合計金額
    excel_sales_drink_total = 0 # 該当月のドリンク売上合計金額
    excel_sales_dinner_visitor = 0 # 該当月のディナー客合計人数
    excel_sales_lunch_visitor = 0 # 該当月のランチ客合計人数
    excel_sales_total_visitor = 0 # 該当月の合計客人数

    """該当月の日報データ作成"""
    Sales_month_data= Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    for s in Sales_month_data:
        day = s.date.strftime("%d")
        dinner_sales = s.total_sales - s.lunch_sales
        food_sales = s.total_sales - s.drink_sales
        if s.drink_sales != 0 and s.total_sales != 0:
            percent = s.drink_sales / s.total_sales * 100
        else:
            percent = 0
        total_guest = s.dinner_guest + s.lunch_guest
        excel_sales_list.append([day, s.total_sales, dinner_sales, s.lunch_sales, food_sales, s.drink_sales, percent, s.dinner_guest, s.lunch_guest, total_guest])

    """Excel出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '日報{0}年{1}月度'.format(year, month)
    ws.cell(row=1, column=1, value="日付")
    ws.cell(row=1, column=2, value="総売上")
    ws.cell(row=1, column=3, value="ディナー")
    ws.cell(row=1, column=4, value="ランチ")
    ws.cell(row=1, column=5, value="フード")
    ws.cell(row=1, column=6, value="ドリンク")
    ws.cell(row=1, column=7, value="ドリンク％")
    ws.cell(row=1, column=8, value="客数(D)")
    ws.cell(row=1, column=9, value="客数(L)")
    ws.cell(row=1, column=10, value="合計入客数")
    for i in range(len(excel_sales_list)):
        ws.cell(row=excel_sales_row, column=1, value=excel_sales_list[i][0])
        ws.cell(row=excel_sales_row, column=2, value=excel_sales_list[i][1])
        ws.cell(row=excel_sales_row, column=3, value=excel_sales_list[i][2])
        ws.cell(row=excel_sales_row, column=4, value=excel_sales_list[i][3])
        ws.cell(row=excel_sales_row, column=5, value=excel_sales_list[i][4])
        ws.cell(row=excel_sales_row, column=6, value=excel_sales_list[i][5])
        ws.cell(row=excel_sales_row, column=7, value=excel_sales_list[i][6])
        ws.cell(row=excel_sales_row, column=8, value=excel_sales_list[i][7])
        ws.cell(row=excel_sales_row, column=9, value=excel_sales_list[i][8])
        ws.cell(row=excel_sales_row, column=10, value=excel_sales_list[i][9])
        excel_sales_row += 1
        excel_sales_sales_total += excel_sales_list[i][1]
        excel_sales_dinner_total += excel_sales_list[i][2]
        excel_sales_lunch_total += excel_sales_list[i][3]
        excel_sales_food_total += excel_sales_list[i][4]
        excel_sales_drink_total += excel_sales_list[i][5]
        excel_sales_dinner_visitor += excel_sales_list[i][7]
        excel_sales_lunch_visitor += excel_sales_list[i][8]
        excel_sales_total_visitor += excel_sales_list[i][9]
    ws.cell(row=excel_sales_row, column=1, value="合計")
    ws.cell(row=excel_sales_row, column=2, value=excel_sales_sales_total)
    ws.cell(row=excel_sales_row, column=3, value=excel_sales_dinner_total)
    ws.cell(row=excel_sales_row, column=4, value=excel_sales_lunch_total)
    ws.cell(row=excel_sales_row, column=5, value=excel_sales_food_total)
    ws.cell(row=excel_sales_row, column=6, value=excel_sales_drink_total)
    ws.cell(row=excel_sales_row, column=7, value=0)
    ws.cell(row=excel_sales_row, column=8, value=excel_sales_dinner_visitor)
    ws.cell(row=excel_sales_row, column=9, value=excel_sales_lunch_visitor)
    ws.cell(row=excel_sales_row, column=10, value=excel_sales_total_visitor)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'sales_{0}_{1}.xlsx'.format(year, month)

    wb.save(response)
    return response

"""減価率出力"""
def makeRateExe(request, year, month):
    excel_sales_total = 0
    excel_sales_drink = 0
    excel_sales_food = 0
    excel_sales_lunch = 0
    excel_sales_dinner = 0
    excel_sales_guest = 0
    excel_sales_average = 0
    excel_sales_day = 0
    excel_cost_total = 0
    excel_cost_food = 0
    excel_cost_food_cash = 0
    excel_cost_food_delivery = 0
    excel_cost_drink = 0
    excel_cost_other = 0
    excel_food_rate = 0
    excel_drink_rate = 0
    fill = PatternFill(fill_type='solid', fgColor='d5e6fc')
    border = Border(
        left=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        right=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        top=Side(
            border_style="thin",
            color="A7A7A7"
        ),
        bottom=Side(
            border_style="thin",
            color="A7A7A7"
        )
    )
    """売上関連データ作成"""
    total_data = Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    for sales in total_data:
        excel_sales_total += sales.total_sales
    total_data = Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    for sales in total_data:
        excel_sales_drink += sales.drink_sales
    excel_sales_food = excel_sales_total - excel_sales_drink
    total_data = Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    for sales in total_data:
        excel_sales_lunch += sales.lunch_sales
    excel_sales_dinner = excel_sales_total - excel_sales_lunch
    total_data = Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    for sales in total_data:
        excel_sales_guest += sales.dinner_guest
        excel_sales_guest += sales.lunch_guest
    total_data = Sales.objects.filter(author=request.user, date__year=year, date__month=month)
    if len(total_data) != 0:
        excel_sales_average = round(excel_sales_total / len(total_data), 2)
        excel_sales_day = len(total_data)

    """経費関連データ作成"""
    total_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month)
    for cost in total_data:
        excel_cost_total += cost.price
    food_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__category__category_name='食材')
    for cost in  food_data:
        excel_cost_food += cost.price
    food_cash_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__category__category_name='食材', shop__flag=0)
    for cost in  food_cash_data:
        excel_cost_food_cash += cost.price
    food_delivery_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__category__category_name='食材', shop__flag=1)
    for cost in  food_delivery_data:
        excel_cost_food_delivery += cost.price
    drink_data = Cost.objects.filter(author=request.user, date__year=year, date__month=month, shop__category__category_name='ドリンク')
    for cost in  drink_data:
        excel_cost_drink += cost.price
    excel_cost_other = excel_cost_total - (excel_cost_food + excel_cost_drink)

    """原価率データ作成"""
    if excel_sales_food != 0:
        excel_food_rate = round(excel_cost_food / excel_sales_food * 100, 2)
    if excel_sales_drink != 0:
        excel_drink_rate = round(excel_cost_drink / excel_sales_drink * 100, 2)

    """Excel出力"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '原価率{0}年{1}月度'.format(year, month)
    sn = ws.cell(row=1, column=1, value="売上")
    sn.fill = fill
    sn.border = border
    ws.cell(row=2, column=1, value="総売上")
    ws.cell(row=2, column=2, value=excel_sales_total)
    ws.cell(row=3, column=1, value="ランチ")
    ws.cell(row=3, column=2, value=excel_sales_lunch)
    ws.cell(row=4, column=1, value="ディナー")
    ws.cell(row=4, column=2, value=excel_sales_dinner)
    ws.cell(row=5, column=1, value="フード")
    ws.cell(row=5, column=2, value=excel_sales_food)
    ws.cell(row=6, column=1, value="ドリンク")
    ws.cell(row=6, column=2, value=excel_sales_drink)
    ws.cell(row=7, column=1, value="入客数")
    ws.cell(row=7, column=2, value=excel_sales_guest)
    ws.cell(row=8, column=1, value="一日平均")
    ws.cell(row=8, column=2, value=excel_sales_average)
    ws.cell(row=9, column=1, value="営業日数")
    ws.cell(row=9, column=2, value=excel_sales_day)
    sn = ws.cell(row=10, column=1, value="経費")
    sn.fill = fill
    sn.border = border
    ws.cell(row=11, column=1, value="総経費")
    ws.cell(row=11, column=2, value=excel_cost_total)
    ws.cell(row=12, column=1, value="フード")
    ws.cell(row=12, column=2, value=excel_cost_food)
    ws.cell(row=13, column=1, value="フード(現金仕入)")
    ws.cell(row=13, column=2, value=excel_cost_food_cash)
    ws.cell(row=14, column=1, value="フード(業者仕入)")
    ws.cell(row=14, column=2, value=excel_cost_food_delivery)
    ws.cell(row=15, column=1, value="ドリンク")
    ws.cell(row=15, column=2, value=excel_cost_drink)
    ws.cell(row=16, column=1, value="その他")
    ws.cell(row=16, column=2, value=excel_cost_other)
    sn = ws.cell(row=17, column=1, value="原価率")
    sn.fill = fill
    sn.border = border
    ws.cell(row=18, column=1, value="フード原価率")
    ws.cell(row=18, column=2, value=excel_food_rate)
    ws.cell(row=19, column=1, value="ドリンク原価率")
    ws.cell(row=19, column=2, value=excel_drink_rate)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'rate_{0}_{1}.xlsx'.format(year, month)

    wb.save(response)
    return response